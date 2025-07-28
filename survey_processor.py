# survey_processor.py
import pandas as pd
from datetime import datetime
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule

# openpyxl is used by pandas for writing .xlsx files, ensure it's in requirements.txt

def read_rid_file_from_stream(file_stream):
    """Reads the RID lookup CSV file from a file stream."""
    try:
        df = pd.read_csv(file_stream)
        df.columns = df.columns.str.strip().str.lower()
        
        # Validate critical columns exist
        if df.empty:
            raise ValueError("RID file appears to be empty. Please check your CSV file.")
        
        # Check for basic expected columns
        expected_cols = ['pid']
        missing_cols = [col for col in expected_cols if col not in df.columns]
        if missing_cols:
            available_cols = ", ".join(df.columns[:5])  # Show first 5 columns
            raise ValueError(f"RID file missing required columns: {missing_cols}. Available columns: {available_cols}...")
            
        return df
    except pd.errors.EmptyDataError:
        raise ValueError("RID file is empty or contains no valid data.")
    except pd.errors.ParserError as e:
        raise ValueError(f"RID file format error: Could not parse CSV file. Please ensure it's a valid CSV format. Details: {str(e)}")
    except UnicodeDecodeError:
        raise ValueError("RID file encoding error: Please ensure the CSV file is saved with UTF-8 encoding.")
    except Exception as e:
        raise ValueError(f"Error reading RID file: {str(e)}")

def read_metrics_file_from_stream(file_stream):
    """Reads the Marketplace Metrics Excel file from a file stream."""
    try:
        df = pd.read_excel(
            file_stream,
            sheet_name='Marketplace Metrics by PID',
            skiprows=5
        )
        df.columns = df.columns.str.strip().str.lower()
        
        # Validate file structure
        if df.empty:
            raise ValueError("PID Metrics file appears to be empty after skipping header rows. Please check your Excel file structure.")
            
        # Check for critical columns
        if 'pid' not in df.columns:
            available_cols = ", ".join(df.columns[:5])  # Show first 5 columns
            raise ValueError(f"PID Metrics file missing 'PID' column. Available columns: {available_cols}...")
            
        if 'pid' in df.columns:
            df['pid'] = df['pid'].astype(str).str.strip()
        return df
    except FileNotFoundError:
        raise ValueError("Could not find the specified sheet 'Marketplace Metrics by PID' in the Excel file.")
    except ValueError as ve:
        if "Worksheet named" in str(ve):
            raise ValueError("Excel file must contain a sheet named 'Marketplace Metrics by PID'. Please check your file format.")
        raise ve  # Re-raise ValueError with original message
    except Exception as e:
        if "xlrd" in str(e) or "openpyxl" in str(e):
            raise ValueError("Excel file format error: Please ensure you're uploading a valid .xlsx file exported from SSRS.")
        raise ValueError(f"Error reading PID Metrics file: {str(e)}")

def get_formatting_ranges(header, data_rows, total_column=None, has_col_total_row=False):
    """
    Helper function to calculate formatting exclusions and last data row for conditional formatting.
    header: list of column names (Excel order, 1-based for openpyxl)
    data_rows: total number of rows written (including totals row)
    total_column: name of total column (e.g. 'Total_Flagged', 'Row Total')
    has_col_total_row: True if the last row is a column total row
    Returns: (exclude_cols, last_data_row)
    """
    exclude_cols = [1]  # Always exclude first column (index 1)
    if '-n/a-' in header:
        exclude_cols.append(header.index('-n/a-') + 1)
    if total_column and total_column in header:
        exclude_cols.append(header.index(total_column) + 1)
    # Only format up to last data row (exclude column total row if present)
    last_data_row = data_rows - 1 if has_col_total_row else data_rows
    return exclude_cols, last_data_row

def apply_conditional_formatting(worksheet, start_col, end_col, data_rows, exclude_cols=None, total_rows=None):
    """Apply conditional formatting to specified range."""
    if exclude_cols is None:
        exclude_cols = []
    # Guard: skip if no data rows
    if data_rows < 2:
        return
    color_scale_rule = ColorScaleRule(
        start_type='num',
        start_value=0,
        start_color='FFFFFF',
        end_type='num',
        end_value=total_rows,
        end_color='f82b1b'
    )
    for col in range(start_col, end_col + 1):
        if col not in exclude_cols:
            col_letter = worksheet.cell(row=1, column=col).column_letter
            cell_range = f"{col_letter}2:{col_letter}{data_rows}"
            # Guard: skip if cell_range is not valid
            if int(data_rows) < 2:
                continue
            try:
                worksheet.conditional_formatting.add(cell_range, color_scale_rule)
            except Exception as e:
                # Log and skip this column if openpyxl fails
                print(f"Conditional formatting skipped for {cell_range}: {e}")
                continue

def apply_na_column_formatting(worksheet, header):
    """Apply dark green formatting to -n/a- column if present."""
    if '-n/a-' in header:
        na_col_idx = header.index('-n/a-') + 1  # openpyxl is 1-based
        dark_green_font = Font(color="006400")
        for row in worksheet.iter_rows(min_row=2, min_col=na_col_idx, max_col=na_col_idx, max_row=worksheet.max_row):
            for cell in row:
                cell.font = dark_green_font
        worksheet.cell(row=1, column=na_col_idx).font = dark_green_font

def format_pivot_sheet(worksheet, header, data_rows, total_column, total_rows, has_col_total_row=True):
    """Standardized pivot formatting: conditional formatting + -n/a- styling."""
    exclude_cols, last_data_row = get_formatting_ranges(
        header, data_rows, total_column=total_column, has_col_total_row=has_col_total_row
    )
    apply_conditional_formatting(
        worksheet,
        start_col=2,
        end_col=len(header),
        data_rows=last_data_row,
        exclude_cols=exclude_cols,
        total_rows=total_rows
    )
    apply_na_column_formatting(worksheet, header)

def add_check_results_pivot(writer, df_merged):
    """Add pivot table showing all check flags by supplier."""
    workbook = writer.book
    check_columns = [
        "Poor_Conv_Rate", "New_User_Bot", "High_Security",
        "Speeder", "High_LOI", "High_RR"
    ]
    
    # Create pivot for check results
    ws_pivot = workbook.create_sheet('Flags Pivot (Multi)')
    
    # Calculate counts by supplier
    supplier_stats = []
    for supplier in df_merged['supplier_bu'].unique():
        supplier_df = df_merged[df_merged['supplier_bu'] == supplier]
        
        # Count rows with any True flag
        has_any_flag = supplier_df[check_columns].any(axis=1)
        total_flagged = has_any_flag.sum()
        
        # Count rows with no flags (true -n/a- count)
        no_flags = ~has_any_flag
        na_count = no_flags.sum()
        
        # Get individual flag counts
        flag_counts = supplier_df[check_columns].sum()
        
        stats = {
            'supplier_bu': supplier,
            '-n/a-': na_count,
            **flag_counts.to_dict(),
            'Total_Flagged': total_flagged
        }
        supplier_stats.append(stats)
    
    pivot_df = pd.DataFrame(supplier_stats)
    pivot_df = pivot_df.sort_values('Total_Flagged', ascending=False)
    
    # Add column totals with proper supplier_bu value
    totals = pd.DataFrame([{
        'supplier_bu': 'Column Total',
        '-n/a-': pivot_df['-n/a-'].sum(),
        **{col: pivot_df[col].sum() for col in check_columns},
        'Total_Flagged': pivot_df['Total_Flagged'].sum()
    }])
    pivot_df = pd.concat([pivot_df, totals], ignore_index=True)
    
    # Arrange columns in desired order:
    # 1. supplier_bu first
    # 2. -n/a- second
    # 3. check columns sorted by total count
    # 4. Total_Flagged last
    check_totals = pivot_df[check_columns].sum()
    sorted_check_cols = sorted(check_columns, key=lambda x: check_totals[x], reverse=True)
    
    header = ['supplier_bu', '-n/a-'] + sorted_check_cols + ['Total_Flagged']
    
    # Write to Excel with formatting
    ws_pivot.append(header)
    
    # Write data rows
    for _, row in pivot_df.iterrows():
        ws_pivot.append([row['supplier_bu']] + [row[col] for col in header[1:]])
    
    # Set supplier_bu column width to 200 pixels
    # Convert pixels to Excel column width units (approximately)
    excel_width = 200 / 7  # Excel width units are roughly 7 pixels
    ws_pivot.column_dimensions['A'].width = excel_width

    # Use helper for exclusions and data range
    exclude_cols, last_data_row = get_formatting_ranges(
        header, len(pivot_df), total_column='Total_Flagged', has_col_total_row=True
    )
    total_rows = len(df_merged)
    apply_conditional_formatting(
        ws_pivot,
        start_col=2,
        end_col=len(header),
        data_rows=last_data_row,
        exclude_cols=exclude_cols,
        total_rows=total_rows
    )
    
    # Apply existing -n/a- column formatting
    dark_green_font = Font(color="006400")
    # Fix: Check if -n/a- column exists and use apply_na_column_formatting instead
    apply_na_column_formatting(ws_pivot, header)

def add_pivot_and_format(writer, df_merged):
    """
    Adds pivot tables to the Excel workbook and applies basic formatting.
    """
    workbook = writer.book
    
    # --- Combined Data Sheet Formatting (from original) ---
    ws_combined = writer.sheets.get('Combined Data') # Get the sheet by name
    if ws_combined:
        ws_combined.auto_filter.ref = ws_combined.dimensions
        ws_combined.freeze_panes = ws_combined['A2']

    # --- Create Pivot Table ---
    if 'supplier_bu' not in df_merged.columns or 'Observation' not in df_merged.columns:
        raise ValueError("Required columns 'supplier_bu' or 'Observation' not found in processed data. This may indicate a problem with the input files.")

    try:
        pivot = (
            df_merged.groupby(['supplier_bu', 'Observation'])
            .size()
            .reset_index(name='Count')
            .pivot(index='supplier_bu', columns='Observation', values='Count')
            .fillna(0)
            .astype(int)
        )

        if pivot.empty:
            raise ValueError("No data available for pivot table generation. Please check that your input files contain valid data.")

        # Add row totals
        pivot['Row Total'] = pivot.sum(axis=1)

        # Add column totals
        col_totals = pivot.sum(axis=0)
        col_totals.name = 'Column Total'
        pivot = pd.concat([pivot, pd.DataFrame([col_totals], index=['Column Total'])])


        # Sort rows by row total (descending), keep total row at the end
        if 'Row Total' in pivot.columns and 'Column Total' in pivot.index:
            # Separate the 'Column Total' row
            total_row_df = pivot.loc[['Column Total']]
            pivot_data_rows = pivot.drop('Column Total')
            
            # Sort the data rows
            pivot_data_rows = pivot_data_rows.sort_values(by='Row Total', ascending=False)
            
            # Concatenate sorted data rows with the total row
            pivot = pd.concat([pivot_data_rows, total_row_df])


        # Sort columns: "-n/a-" first, then by column total (descending), then Row Total last
        cols = list(pivot.columns)
        # Use .get(c, 0) for robustness if a column name is unexpectedly missing from col_totals
        col_total_values = pivot.loc['Column Total'] if 'Column Total' in pivot.index else pivot.sum(axis=0)

        obs_cols = [c for c in cols if c not in ['Row Total']]
        sorted_obs_cols = []

        if '-n/a-' in obs_cols:
            sorted_obs_cols.append('-n/a-')
            obs_cols_no_na = [c for c in obs_cols if c != '-n/a-']
        else:
            obs_cols_no_na = list(obs_cols)
            
        # Sort remaining observation columns by their total count
        obs_cols_sorted = sorted(obs_cols_no_na, key=lambda c: col_total_values.get(c, 0), reverse=True)
        sorted_obs_cols.extend(obs_cols_sorted)

        if 'Row Total' in cols:
            sorted_obs_cols.append('Row Total')
        
        pivot = pivot[sorted_obs_cols]

        # Write pivot table to new sheet
        ws_pivot = workbook.create_sheet('Flags Pivot (Priority)')
        
        # Write header (supplier_bu and then pivot columns)
        header = ['supplier_bu'] + list(pivot.columns)
        ws_pivot.append(header)
        
        # Write data rows
        for supplier_bu_index, row_data in pivot.iterrows():
            ws_pivot.append([supplier_bu_index] + list(row_data.values))

        # Set supplier_bu column width to 200 pixels
        excel_width = 200 / 7  # Excel width units are roughly 7 pixels
        ws_pivot.column_dimensions['A'].width = excel_width

        exclude_cols, last_data_row = get_formatting_ranges(
            header, len(pivot), total_column='Row Total', has_col_total_row=True
        )
        total_rows = len(df_merged)
        apply_conditional_formatting(
            ws_pivot,
            start_col=2,
            end_col=len(header),
            data_rows=last_data_row,
            exclude_cols=exclude_cols,
            total_rows=total_rows
        )

        # --- Style "-n/a-" column in dark green ---
        dark_green_font = Font(color="006400")  # Hex for dark green

        # Find the column index for "-n/a-"
        try:
            na_col_idx = header.index('-n/a-') + 1  # openpyxl is 1-based
            for row in ws_pivot.iter_rows(min_row=2, min_col=na_col_idx, max_col=na_col_idx, max_row=ws_pivot.max_row):
                for cell in row:
                    cell.font = dark_green_font
            # Also style the header cell
            ws_pivot.cell(row=1, column=na_col_idx).font = dark_green_font
        except ValueError:
            pass  # "-n/a-" column not present

        # --- Additional Pivots ---

        # Helper: get just the date part from entrydate
        df_merged['entrydate_only'] = df_merged['entrydate'].astype(str).str.split('T').str[0]

        # 1. Pivot: entrydate_only vs supplier_bu
        if 'entrydate_only' in df_merged.columns and 'supplier_bu' in df_merged.columns:
            pivot_entrydate_supplier = (
                df_merged.groupby(['entrydate_only', 'supplier_bu'])
                .size()
                .reset_index(name='Count')
                .pivot(index='entrydate_only', columns='supplier_bu', values='Count')
                .fillna(0)
                .astype(int)
            )
            
            # Sort index by date (ascending)
            pivot_entrydate_supplier = pivot_entrydate_supplier.sort_index()
            
            # Add row totals
            pivot_entrydate_supplier['Row Total'] = pivot_entrydate_supplier.sum(axis=1)
            
            # Add column totals
            col_totals = pivot_entrydate_supplier.sum()
            col_totals.name = 'Column Total'
            pivot_entrydate_supplier = pd.concat([pivot_entrydate_supplier, pd.DataFrame([col_totals], index=['Column Total'])])
            
            # Sort columns by total count (highest first) but keep Row Total last
            cols = list(pivot_entrydate_supplier.columns)
            if 'Row Total' in cols:
                cols.remove('Row Total')
            col_totals_sorted = pivot_entrydate_supplier.loc[pivot_entrydate_supplier.index != 'Column Total', cols].sum()
            sorted_cols = col_totals_sorted.sort_values(ascending=False).index
            sorted_cols = list(sorted_cols) + ['Row Total']
            pivot_entrydate_supplier = pivot_entrydate_supplier[sorted_cols]

            ws_pivot1 = workbook.create_sheet('Pivot EntryDate x Supplier')
            header = ['entrydate'] + list(pivot_entrydate_supplier.columns)
            ws_pivot1.append(header)
            for idx, row_data in pivot_entrydate_supplier.iterrows():
                ws_pivot1.append([idx] + list(row_data.values))
            # Set all columns to 100 pixels width
            excel_width = 100 / 7  # Excel width units are roughly 7 pixels
            for col_idx in range(1, len(header) + 1):
                col_letter = ws_pivot1.cell(row=1, column=col_idx).column_letter
                ws_pivot1.column_dimensions[col_letter].width = excel_width

            format_pivot_sheet(ws_pivot1, header, len(pivot_entrydate_supplier), 'Row Total', len(df_merged))

        # 2. Pivot: entrydate_only vs Observation
        if 'entrydate_only' in df_merged.columns and 'Observation' in df_merged.columns:
            pivot_entrydate_obs = (
                df_merged.groupby(['entrydate_only', 'Observation'])
                .size()
                .reset_index(name='Count')
                .pivot(index='entrydate_only', columns='Observation', values='Count')
                .fillna(0)
                .astype(int)
            )
            
            # Sort index by date (ascending)
            pivot_entrydate_obs = pivot_entrydate_obs.sort_index()
            
            # Add row totals
            pivot_entrydate_obs['Row Total'] = pivot_entrydate_obs.sum(axis=1)
            
            # Add column totals
            col_totals = pivot_entrydate_obs.sum()
            col_totals.name = 'Column Total'
            pivot_entrydate_obs = pd.concat([pivot_entrydate_obs, pd.DataFrame([col_totals], index=['Column Total'])])
            
            # Sort columns: "-n/a-" first, then others by total (highest first), then Row Total last
            cols = list(pivot_entrydate_obs.columns)
            sorted_cols = []
            if '-n/a-' in cols:
                sorted_cols.append('-n/a-')
                cols.remove('-n/a-')
            if 'Row Total' in cols:
                cols.remove('Row Total')
            col_totals_sorted = pivot_entrydate_obs.loc[pivot_entrydate_obs.index != 'Column Total', cols].sum()
            remaining_cols = col_totals_sorted.sort_values(ascending=False).index
            sorted_cols.extend(remaining_cols)
            sorted_cols.append('Row Total')
            pivot_entrydate_obs = pivot_entrydate_obs[sorted_cols]

            ws_pivot2 = workbook.create_sheet('Pivot EntryDate x Flags')
            header = ['entrydate'] + list(pivot_entrydate_obs.columns)
            ws_pivot2.append(header)
            for idx, row in pivot_entrydate_obs.iterrows():
                ws_pivot2.append([idx] + list(row.values))
            ws_pivot2.column_dimensions['A'].width = 100 / 7

            format_pivot_sheet(ws_pivot2, header, len(pivot_entrydate_obs), 'Row Total', len(df_merged))

            # Format -n/a- column in dark green (like other pivots)
            try:
                na_col_idx = header.index('-n/a-') + 1  # openpyxl is 1-based
                dark_green_font = Font(color="006400")
                for row in ws_pivot2.iter_rows(min_row=2, min_col=na_col_idx, max_col=na_col_idx, max_row=ws_pivot2.max_row):
                    for cell in row:
                        cell.font = dark_green_font
                ws_pivot2.cell(row=1, column=na_col_idx).font = dark_green_font
            except ValueError:
                pass  # "-n/a-" column not present

    except Exception as e:
        print(f"Error creating pivot table: {e}")
        # Return early to prevent further errors
        return

def generate_survey_report(
    rid_file_stream, metrics_file_stream, actual_loi, output_dir,
    conversion_rate_threshold=10,
    security_terms_threshold=30,
    speeder_multiplier=3,
    high_loi_multiplier=3,
    negative_recs_rate_threshold=15,
    process_status_26_only=True,
    use_datetime_for_newuser=True
):
    """
    Processes the survey files and generates an Excel report with observations and a pivot table.
    """
    try:
        if not (3 <= actual_loi <= 100):
            raise ValueError("Survey Actual LOI must be between 3 and 100.")

        # Read files with enhanced error handling
        try:
            rid_df = read_rid_file_from_stream(rid_file_stream)
        except Exception as e:
            raise ValueError(f"RID file error: {str(e)}")
            
        try:
            metrics_df = read_metrics_file_from_stream(metrics_file_stream)
        except Exception as e:
            raise ValueError(f"PID Metrics file error: {str(e)}")

        if 'pid' not in rid_df.columns or 'pid' not in metrics_df.columns:
            raise ValueError("Critical Error: 'pid' column not found in one or both input files. Please check column headers (case-sensitive: should be 'PID' or 'pid').")
        
        # Enhanced merge validation
        rid_pids = set(rid_df['pid'].astype(str).str.strip())
        metrics_pids = set(metrics_df['pid'].astype(str).str.strip())
        
        if len(rid_pids & metrics_pids) == 0:
            raise ValueError("No matching PIDs found between RID and Metrics files. Please ensure both files contain the same PIDs.")
        
        # Ensure 'pid' columns are strings for reliable merging
        rid_df['pid'] = rid_df['pid'].astype(str).str.strip()
        metrics_df['pid'] = metrics_df['pid'].astype(str).str.strip()

        merged_df = pd.merge(
            rid_df,
            metrics_df,
            on='pid',
            how='left',
            indicator=False
        )

        # Filter for status=26 if requested
        if process_status_26_only and 'client_responsestatusid' in merged_df.columns:
            merged_df = merged_df[merged_df['client_responsestatusid'].astype(str) == '26'].copy()

        # Scale system_conversion_rate and negative_recs_rate to 0-100 immediately after merging
        if "system_conversion_rate" in merged_df.columns:
            merged_df["system_conversion_rate"] = merged_df["system_conversion_rate"] * 100
        if "negative_recs_rate" in merged_df.columns:
            merged_df["negative_recs_rate"] = merged_df["negative_recs_rate"] * 100

        # Apply observation logic and add check columns
        merged_df = apply_pid_observation_logic(
            merged_df,
            actual_loi=actual_loi,
            conversion_rate_threshold=conversion_rate_threshold,
            security_terms_threshold=security_terms_threshold,
            speeder_multiplier=speeder_multiplier,
            high_loi_multiplier=high_loi_multiplier,
            negative_recs_rate_threshold=negative_recs_rate_threshold,
            session_loi_checks=True,
            use_datetime_for_newuser=use_datetime_for_newuser
        )

        # --- Insert High_Security criteria % calculation column ---
        high_security_pct = (
            merged_df["sum_f_and_g_column"] / merged_df["total_system_entrants"]
        ).fillna(0) * 100
        high_security_pct = high_security_pct.round(2)
        cols = list(merged_df.columns)
        if "system_conversion_rate" in cols:
            idx = cols.index("system_conversion_rate") + 1
            merged_df.insert(idx, "High_Security_Pct", high_security_pct)
        else:
            merged_df["High_Security_Pct"] = high_security_pct

        # --- Insert first/last entry date match column ---
        match_col = (merged_df["first_entry_date"] == merged_df["last_entry_date"])
        if "last_entry_date" in merged_df.columns:
            cols = list(merged_df.columns)  # Update cols list
            idx = cols.index("last_entry_date") + 1
            merged_df.insert(idx, "FirstLastDateMatch", match_col)
        else:
            merged_df["FirstLastDateMatch"] = match_col

        # --- Insert Diff Days column after last_entry_date ---
        if "first_entry_date" in merged_df.columns and "last_entry_date" in merged_df.columns:
            # Convert to datetime if not already
            merged_df["first_entry_date_dt"] = pd.to_datetime(merged_df["first_entry_date"], errors='coerce')
            merged_df["last_entry_date_dt"] = pd.to_datetime(merged_df["last_entry_date"], errors='coerce')
            diff_days = (merged_df["last_entry_date_dt"] - merged_df["first_entry_date_dt"]).dt.days
            # Insert after last_entry_date
            cols = list(merged_df.columns)
            idx = cols.index("last_entry_date") + 1
            merged_df.insert(idx, "Diff Days", diff_days)
            # Remove temp columns
            merged_df.drop(["first_entry_date_dt", "last_entry_date_dt"], axis=1, inplace=True)
        else:
            merged_df["Diff Days"] = None

        # --- Round columns to 2 decimal places ---
        if "system_conversion_rate" in merged_df.columns:
            merged_df["system_conversion_rate"] = merged_df["system_conversion_rate"].round(2)
        if "negative_recs_rate" in merged_df.columns:
            merged_df["negative_recs_rate"] = merged_df["negative_recs_rate"].round(2)
        if "High_Security_Pct" in merged_df.columns:
            merged_df["High_Security_Pct"] = merged_df["High_Security_Pct"].round(2)

        # --- Add Flag_Count column at the end ---
        flag_columns = [
            "Poor_Conv_Rate", "New_User_Bot", "High_Security",
            "Speeder", "High_LOI", "High_RR"
        ]
        merged_df["Flag_Count"] = merged_df[flag_columns].sum(axis=1)

        # Remove blank columns (all values are NaN or empty) before writing to Excel
        merged_df = merged_df.dropna(axis=1, how='all')

        # --- Generate Excel File ---
        os.makedirs(output_dir, exist_ok=True) # Ensure output directory exists
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"RID-PID_Report_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                merged_df.to_excel(writer, sheet_name='Combined Data', index=False)
                add_pivot_and_format(writer, merged_df)     # Creates "Flags Pivot (Priority)" and time-series pivots
                add_check_results_pivot(writer, merged_df)  # Creates "Flags Pivot (Multi)"

                # Reorder sheets to match desired sequence
                wb = writer.book
                desired_order = [
                    "Combined Data",
                    "Flags Pivot (Priority)", 
                    "Flags Pivot (Multi)",
                    "Pivot EntryDate x Supplier",
                    "Pivot EntryDate x Flags"
                ]
                
                # Reorder existing sheets
                existing_sheets = []
                for sheet_name in desired_order:
                    if sheet_name in wb.sheetnames:
                        existing_sheets.append(wb[sheet_name])
                
                # Remove all sheets from workbook
                wb._sheets.clear()
                
                # Add sheets back in desired order
                for sheet in existing_sheets:
                    wb._sheets.append(sheet)

                # --- Create DenyList_Draft sheet ---
                deny_cols = [
                    "pid", "supplierid", "name", "Observation",
                    "Poor_Conv_Rate", "New_User_Bot", "High_Security",
                    "Speeder", "High_LOI", "High_RR", "Flag_Count",
                    "Diff Days"  # <-- ensure this is included for DenyList_Draft
                ]
                # Only keep columns that exist in merged_df
                deny_cols_present = [c for c in deny_cols if c in merged_df.columns]
                deny_df = merged_df[deny_cols_present].copy()

                # Rename columns first, then insert Deny Criteria
                col_map = {
                    "pid": "PID",
                    "supplierid": "Supplier ID", 
                    "name": "Supplier Name"
                }
                deny_df = deny_df.rename(columns=col_map)
                
                # Insert "Deny Criteria" column after "Supplier Name"
                cols_list = list(deny_df.columns)
                if "Supplier Name" in cols_list:
                    name_idx = cols_list.index("Supplier Name")
                    deny_df.insert(name_idx + 1, "Deny Criteria", 10)
                else:
                    # Fallback: insert after second column
                    deny_df.insert(2, "Deny Criteria", 10)

                # Create the sheet and write data properly
                deny_sheet = wb.create_sheet("DenyList_Draft")
                
                # Write headers
                for col_idx, col_name in enumerate(deny_df.columns, 1):
                    deny_sheet.cell(row=1, column=col_idx, value=col_name)
                
                # Write data rows
                for row_idx, row_data in enumerate(deny_df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row_data, 1):
                        deny_sheet.cell(row=row_idx, column=col_idx, value=value)

                # Enable auto-filter and freeze first row
                deny_sheet.auto_filter.ref = deny_sheet.dimensions
                deny_sheet.freeze_panes = deny_sheet['A2']
                # Set header alignment to left for Combined Data and DenyList_Draft
                from openpyxl.styles import Alignment
                left_align = Alignment(horizontal='left')
                # Combined Data
                combined_sheet = wb["Combined Data"]
                for cell in combined_sheet[1]:
                    cell.alignment = left_align
                # DenyList_Draft
                for cell in deny_sheet[1]:
                    cell.alignment = left_align
                # --- Enhanced Conditional Formatting for Combined Data ---
                from openpyxl.formatting.rule import ColorScaleRule
                from openpyxl.styles import Font
                import numpy as np

                header = [cell.value for cell in combined_sheet[1]]
                n_rows = combined_sheet.max_row

                # Helper to get min/max and ensure numeric
                def get_col_min_max(col_name):
                    if col_name in merged_df.columns:
                        col_data = pd.to_numeric(merged_df[col_name], errors='coerce')
                        col_min = np.nanmin(col_data)
                        col_max = np.nanmax(col_data)
                        return col_min, col_max
                    return None, None

                # Conditional formatting rules for each column
                format_specs = {
                    # system_conversion_rate: high is good, red for low (bad), white for high (good), scale 0-100
                    "system_conversion_rate": {
                        "min_color": "FFFFFF", "max_color": "f82b1b", "min": 0, "max": 100, "reverse": True
                    },
                    # High_Security_Pct: high is bad, red for high (bad), white for low (good), scale 0-100
                    "High_Security_Pct": {
                        "min_color": "FFFFFF", "max_color": "f82b1b", "min": 0, "max": 100, "reverse": False
                    },
                    # negative_recs_rate: high is bad, red for high (bad), white for low (good), scale 0-100
                    "negative_recs_rate": {
                        "min_color": "FFFFFF", "max_color": "f82b1b", "min": 0, "max": 100, "reverse": False
                    },
                    # Flag_Count: fixed scale 0-5, white for low, red for high
                    "Flag_Count": {
                        "min_color": "FFFFFF", "max_color": "f82b1b", "min": 0, "max": 5, "reverse": False
                    },
                    # client_responsestatusid: orange to pickle green
                    "client_responsestatusid": {
                        "min_color": "FFA500", "max_color": "4f9e4f", "reverse": False
                    },
                    # session_loi: 3-color scale yellow-white-yellow
                    "session_loi": {
                        "min_color": "FFFF00", "mid_color": "FFFFFF", "max_color": "FFFF00", "reverse": False, "three_color": True
                    },
                    # supplier_bu_id: sky blue to gray
                    "supplier_bu_id": {
                        "min_color": "87CEEB", "max_color": "808080", "reverse": False
                    },
                    # survey_ccpi: white to yellow
                    "survey_ccpi": {
                        "min_color": "FFFFFF", "max_color": "FFFF00", "reverse": False
                    },
                    # survey_qcpi: white to yellow
                    "survey_qcpi": {
                        "min_color": "FFFFFF", "max_color": "FFFF00", "reverse": False
                    },
                    # Diff Days: white to yellow
                    "Diff Days": {
                        "min_color": "FFFFFF", "max_color": "FFFF00", "reverse": False
                    },
                    # Total count columns: white (0) to yellow (high values)
                    "total_system_entrants": {
                        "min_color": "FFFFFF", "max_color": "FFFF00", "reverse": False
                    },
                    "total_surveys_entered": {
                        "min_color": "FFFFFF", "max_color": "FFFF00", "reverse": False
                    },
                    "total_completes": {
                        "min_color": "FFFFFF", "max_color": "FFFF00", "reverse": False
                    },
                    "total_negative_recs": {
                        "min_color": "FFFFFF", "max_color": "FFFF00", "reverse": False
                    },
                    "total_security_terms_on_marketplace_side": {
                        "min_color": "FFFFFF", "max_color": "FFFF00", "reverse": False
                    },
                    "total_security_terms_on_client_side": {
                        "min_color": "FFFFFF", "max_color": "FFFF00", "reverse": False
                    },
                    "sum_f_and_g_column": {
                        "min_color": "FFFFFF", "max_color": "FFFF00", "reverse": False
                    }
                }

                for col_name, spec in format_specs.items():
                    if col_name in header:
                        col_idx = header.index(col_name) + 1
                        col_letter = combined_sheet.cell(row=1, column=col_idx).column_letter
                        cell_range = f"{col_letter}2:{col_letter}{n_rows}"
                        
                        # Get min/max
                        if "min" in spec and "max" in spec:
                            col_min, col_max = spec["min"], spec["max"]
                        else:
                            col_min, col_max = get_col_min_max(col_name)
                            if col_min is None or col_max is None or col_min == col_max:
                                continue
                        # 3-color scale for session_loi
                        if spec.get("three_color"):
                            col_median = np.nanmedian(pd.to_numeric(merged_df[col_name], errors='coerce'))
                            color_rule = ColorScaleRule(
                                start_type='num', start_value=col_min, start_color=spec["min_color"],
                                mid_type='num', mid_value=col_median, mid_color=spec["mid_color"],
                                end_type='num', end_value=col_max, end_color=spec["max_color"]
                            )
                        else:
                            # Special case: High_Security_Pct should be red for high, white for low
                            if col_name == "High_Security_Pct":
                                color_rule = ColorScaleRule(
                                    start_type='num', start_value=col_min, start_color="FFFFFF",
                                    end_type='num', end_value=col_max, end_color="f82b1b"
                                )
                            elif spec.get("reverse"):
                                color_rule = ColorScaleRule(
                                    start_type='num', start_value=col_min, start_color=spec["max_color"],
                                    end_type='num', end_value=col_max, end_color=spec["min_color"]
                                )
                            else:
                                color_rule = ColorScaleRule(
                                    start_type='num', start_value=col_min, start_color=spec["min_color"],
                                    end_type='num', end_value=col_max, end_color=spec["max_color"]
                                )
                        try:
                            combined_sheet.conditional_formatting.add(cell_range, color_rule)
                        except (TypeError, AttributeError) as e:
                            # Graceful degradation for openpyxl compatibility issues
                            print(f"Warning: Could not apply conditional formatting for {col_name}: {e}")
                            continue

                # Style -n/a- column in dark green if present
                if "-n/a-" in header:
                    na_col_idx = header.index("-n/a-") + 1
                    dark_green_font = Font(color="006400")
                    for row in combined_sheet.iter_rows(min_row=2, min_col=na_col_idx, max_col=na_col_idx, max_row=combined_sheet.max_row):
                        for cell in row:
                            cell.font = dark_green_font
                    combined_sheet.cell(row=1, column=na_col_idx).font = dark_green_font

                # Style Observation column "-n/a-" values in dark green
                if "Observation" in header:
                    obs_col_idx = header.index("Observation") + 1
                    dark_green_font = Font(color="006400")
                    for row in combined_sheet.iter_rows(min_row=2, min_col=obs_col_idx, max_col=obs_col_idx, max_row=combined_sheet.max_row):
                        for cell in row:
                            if str(cell.value) == "-n/a-":
                                cell.font = dark_green_font

        except PermissionError:
            raise ValueError(f"Cannot write to output file. Please ensure the file is not open in Excel: {output_filename}")
        except Exception as e:
            if "openpyxl" in str(e):
                raise ValueError(f"Excel generation error: {str(e)}. The data was processed but Excel formatting failed.")
            raise ValueError(f"Error creating Excel report: {str(e)}")

        return output_path
        
    except ValueError:
        raise  # Re-raise ValueError as-is
    except Exception as e:
        raise ValueError(f"Unexpected error during report generation: {str(e)}")

def generate_pid_only_report(
    metrics_file_stream,
    output_dir,
    conversion_rate_threshold=10,
    security_terms_threshold=30,
    negative_recs_rate_threshold=15,
    use_datetime_for_newuser=True
):
    """
    Processes only the PID Metrics file and generates an Excel report with observations (PID-only mode).
    """
    try:
        # Enhanced file reading with error handling
        try:
            metrics_df = pd.read_excel(metrics_file_stream, sheet_name='Marketplace Metrics by PID', skiprows=5)
        except Exception as e:
            if "Worksheet named" in str(e):
                raise ValueError("Excel file must contain a sheet named 'Marketplace Metrics by PID'. Please check your file format.")
            raise ValueError(f"Error reading PID Metrics file: {str(e)}")
            
        metrics_df.columns = metrics_df.columns.str.strip().str.lower()
        
        if 'pid' not in metrics_df.columns:
            available_cols = ", ".join(metrics_df.columns[:5])
            raise ValueError(f"Critical Error: 'pid' column not found in PID Metrics file. Available columns: {available_cols}...")
            
        if metrics_df.empty:
            raise ValueError("PID Metrics file contains no data rows. Please check your Excel file.")
        
        # Pass None for speeder_multiplier and high_loi_multiplier, and session_loi_checks=False
        metrics_df = apply_pid_observation_logic(
            metrics_df,
            actual_loi=None,
            conversion_rate_threshold=conversion_rate_threshold,
            security_terms_threshold=security_terms_threshold,
            speeder_multiplier=None,
            high_loi_multiplier=None,
            negative_recs_rate_threshold=negative_recs_rate_threshold,
            session_loi_checks=False,
            use_datetime_for_newuser=use_datetime_for_newuser
        )
        
        # --- Generate Excel File ---
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"pid_metrics_report_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            metrics_df.to_excel(writer, sheet_name='PID Metrics Data', index=False)
            workbook = writer.book
            ws_data = writer.sheets['PID Metrics Data']
            ws_data.freeze_panes = ws_data['B2']
            ws_data.auto_filter.ref = ws_data.dimensions
            # Add Observation Pivot with formatting and conditional totals
            if 'Observation' in metrics_df.columns:
                obs_counts = metrics_df['Observation'].value_counts().reset_index()
                obs_counts.columns = ['Observation', 'Count']
                na_row = obs_counts[obs_counts['Observation'] == '-n/a-']
                other_rows = obs_counts[obs_counts['Observation'] != '-n/a-'].sort_values('Count', ascending=False)
                obs_counts_sorted = pd.concat([na_row, other_rows], ignore_index=True)
                # Only add totals if there are multiple unique observations
                add_totals = len(obs_counts_sorted) > 1
                # Do NOT add row total column in PID-only mode
                obs_counts_sorted.to_excel(writer, sheet_name='Observation Pivot', index=False)
                ws_pivot = writer.sheets['Observation Pivot']
                ws_pivot.auto_filter.ref = ws_pivot.dimensions
                ws_pivot.freeze_panes = ws_pivot['B2']
                dark_green_font = Font(color="006400")
                for row in ws_pivot.iter_rows(min_row=2, max_row=2, min_col=1, max_col=2):
                    for cell in row:
                        if cell.value == '-n/a-' or (cell.row == 2 and ws_pivot['A2'].value == '-n/a-'):
                            cell.font = dark_green_font
                # Style Row Total row bold and add column total if needed
                if add_totals:
                    total_row_idx = ws_pivot.max_row
                    # Only add bold if there is a Row Total row (but not a row total column)
                    if ws_pivot.cell(row=total_row_idx, column=1).value == 'Row Total':
                        for cell in ws_pivot[total_row_idx]:
                            cell.font = Font(bold=True)
                    # Add column total (sum of Count column) in the cell below last row
                    ws_pivot.cell(row=ws_pivot.max_row+1, column=1, value='Column Total')
                    ws_pivot.cell(row=ws_pivot.max_row, column=2, value=obs_counts_sorted['Count'].sum())
                    ws_pivot.cell(row=ws_pivot.max_row, column=2).font = Font(bold=True)
    except ValueError:
        raise  # Re-raise ValueError as-is
    except Exception as e:
        raise ValueError(f"Unexpected error in PID-only processing: {str(e)}")

def apply_pid_observation_logic(
    df,
    actual_loi=None,
    conversion_rate_threshold=10,
    security_terms_threshold=30,
    speeder_multiplier=3,
    high_loi_multiplier=3,
    negative_recs_rate_threshold=15,
    session_loi_checks=True,
    use_datetime_for_newuser=True
):
    """
    Applies all non-RID-based observation logic to the DataFrame in-place.
    If session_loi_checks is False, skips Speeder and High LOI checks.
    """
    import pandas as pd
    
    # Validate required columns exist
    required_base_cols = ['system_conversion_rate', 'negative_recs_rate']
    missing_cols = [col for col in required_base_cols if col not in df.columns]
    if missing_cols:
        available_cols = ", ".join(df.columns[:10])
        raise ValueError(f"Missing required columns for analysis: {missing_cols}. Available columns: {available_cols}...")
    
    # Check for datetime columns
    datetime_cols = ['first_entry_date_time', 'last_entry_date_time']
    missing_datetime = [col for col in datetime_cols if col not in df.columns]
    if missing_datetime and use_datetime_for_newuser:
        raise ValueError(f"Missing datetime columns: {missing_datetime}. Please ensure your PID Metrics file contains these columns or switch to date-only mode.")
    
    # Validate security analysis columns
    security_cols = ['sum_f_and_g_column', 'total_system_entrants']
    missing_security = [col for col in security_cols if col not in df.columns]
    if missing_security:
        raise ValueError(f"Missing security analysis columns: {missing_security}. Please ensure your PID Metrics file is complete.")
    
    # Check for session LOI column if needed
    if session_loi_checks and 'session_loi' not in df.columns:
        raise ValueError("Missing 'session_loi' column required for Speeder and High LOI analysis. Please ensure your files are merged correctly.")
    
    try:
        # Dates: Convert to datetime objects if they are not already.
        df["first_entry_date_time"] = pd.to_datetime(df.get("first_entry_date_time"), errors='coerce')
        df["last_entry_date_time"] = pd.to_datetime(df.get("last_entry_date_time"), errors='coerce')
    except Exception as e:
        raise ValueError(f"Error processing datetime columns: {str(e)}. Please check your date/time data format.")
    
    # Set default value
    df["Observation"] = "-n/a-"
    
    # Initialize check columns as False (not blank)
    check_columns = [
        "Poor_Conv_Rate", "New_User_Bot", "High_Security",
        "Speeder", "High_LOI", "High_RR"
    ]
    
    for col in check_columns:
        df[col] = False
    
    # 1. Poor Conversion Rate (0-100 scale)
    mask_poor_conversion = df["system_conversion_rate"] < conversion_rate_threshold
    df.loc[mask_poor_conversion.fillna(False), "Poor_Conv_Rate"] = True
    df.loc[mask_poor_conversion.fillna(False), "Observation"] = "Poor Conversion Rate"

    # 2. New User (bot?)
    if use_datetime_for_newuser:
        mask_new_user = (df["first_entry_date_time"] == df["last_entry_date_time"]) & \
                       (df["first_entry_date_time"].notna()) & \
                       (df["last_entry_date_time"].notna())
    else:
        mask_new_user = (df["first_entry_date"] == df["last_entry_date"]) & \
                       (df["first_entry_date"].notna()) & \
                       (df["last_entry_date"].notna())
    df.loc[mask_new_user.fillna(False), "New_User_Bot"] = True
    df.loc[mask_new_user.fillna(False), "Observation"] = "New User (bot?)"

    # 3. High Security Terms (percentage, so multiply by 100)
    mask_high_security = (df["total_system_entrants"] > 0) & \
                       ((df["sum_f_and_g_column"] / df["total_system_entrants"] * 100) > security_terms_threshold)
    df.loc[mask_high_security.fillna(False), "High_Security"] = True
    df.loc[mask_high_security.fillna(False), "Observation"] = "High Security Terms"

    # 4. Speeder & 5. High LOI
    if session_loi_checks and actual_loi is not None and speeder_multiplier and high_loi_multiplier:
        mask_speeder = df["session_loi"] < (actual_loi / speeder_multiplier)
        df.loc[mask_speeder.fillna(False), "Speeder"] = True
        df.loc[mask_speeder.fillna(False), "Observation"] = "Speeder"

        mask_high_loi = df["session_loi"] > (actual_loi * high_loi_multiplier)
        df.loc[mask_high_loi.fillna(False), "High_LOI"] = True
        df.loc[mask_high_loi.fillna(False), "Observation"] = "High LOI, Distracted"

    # 6. High RR% (0-100 scale)
    mask_high_rr = df["negative_recs_rate"] > negative_recs_rate_threshold
    df.loc[mask_high_rr.fillna(False), "High_RR"] = True
    df.loc[mask_high_rr.fillna(False), "Observation"] = "High RR%"
    
    return df